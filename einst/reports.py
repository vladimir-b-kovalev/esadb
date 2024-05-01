
from pathlib import Path
import os

from django.shortcuts import render, redirect
from django.views.decorators.http import require_http_methods
from django.conf import settings

from datetime import date
from datetime import datetime

import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill, Border, Side
from openpyxl.utils.cell import get_column_letter
from openpyxl.drawing.image import Image    

from utils.reports import ExcelReport, HSTYLE
from utils.files import filename_normal

from mic.models import MIC
from einst.models import EInst
from docstore.models import DocStore

@require_http_methods(['GET'])
def einst_report(request):
    '''
    формирование отчета об электроустановке в многостраничный файл Excel
    '''    
# содание книги        
    projectname = request.session.get('projectname')
    einstid = request.session.get('einstid')
    if einstid != None:
        einst = EInst.objects.get(id = einstid)
# создание отчета
        wb = ExcelReport()
        ws = wb.active
        ws.title = 'ОБЛОЖКА'
        logo = Image(os.path.join(settings.STATIC_ROOT, 'images', 'esa_logo.jpg'))
        logo.width = 600
        logo.height = 100
        ws.add_image(logo, anchor = 'A1')
        ws['A10'] = 'АГЕНТСТВО ЭНЕРГЕТИЧЕСКИХ РЕШЕНИЙ'
        ws['A10'].font = HSTYLE['font']
        ws['A10'].border = HSTYLE['border']
        ws['A10'].fill = HSTYLE['fill']
        ws['A10'].alignment = HSTYLE['alignment']
        ws.column_dimensions['A'].width = 90
# титульный лист
        ws = wb.worksheet_as_page(mdlexample = einst, name = 'ТИТУЛ', note = 'примечание')
# лист - перечень ИИК        
        ws = wb.worksheet_as_list(mdlset = einst.mic_set.all(), name = 'ИИК', note = 'примечание')
# лист - измерительные трансформаторы            
        ws = wb.worksheet_as_2list(mdlset = einst.mic_set.all(), relname = 'ttnexample', name = 'ТТН', note = 'примечание')
# лист - счетчики                
        ws = wb.worksheet_as_2list(mdlset = einst.mic_set.all(), relname = 'meter', name = 'СЧЕТЧИКИ', note = 'примечание')
# лист - каналы связи               
        ws = wb.worksheet_as_list(mdlset = einst.channels.all(), name = 'СВЯЗЬ', note = 'примечание')
# лист - документы                
        ws = wb.worksheet_as_list(mdlset = einst.docs.all(), name = 'ДОКУМЕНТЫ', note = 'примечание')
# лист - контакты                
        ws = wb.worksheet_as_list(mdlset = einst.contacts.all(), name = 'КОНТАКТЫ', note = 'примечание')
# формирование полного имени файла отчета                
        basedir = Path(__file__).resolve().parent.parent
        wbname = filename_normal(einst.name +'_' + date.today().strftime('%m-%d-%y'))
        try: 
# запись файла и внесение в БД        
            wb.save(os.path.join(settings.MEDIA_ROOT, 'docstore', wbname + '.xlsx'))
            report = DocStore()
            report.doctype = 'отчет'
            report.name = 'Сводный отчет об объекте: ' +  einst.name
            report.number = '_'
            report.date = date.today()
            report.docfile = os.path.join('docstore', wbname + '.xlsx')
            report.save()
            einst.docs.add(report)
        except Exception as e: print('EXCEPTION:', e)
    return redirect('../')

