from django.db import models
import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill, Border, Side
from openpyxl.utils.cell import get_column_letter

# стили по умолчанию)
HSTYLE = {'font': Font(name = 'Arial', size = 12),
    'border': Border(
        left = Side(border_style = 'thin', color = 'FF000000'),
        right = Side(border_style = 'thin', color = 'FF000000'),
        top = Side(border_style = 'thin', color = 'FF000000'),
        bottom = Side(border_style = 'thin', color = 'FF000000')),
    'fill': PatternFill(fill_type = 'solid', fgColor = '00C0C0C0'),
    'alignment': Alignment(horizontal = 'center', vertical = 'center')}
    
BSTYLE = {'font': Font(name = 'Arial', size = 12),
    'border': Border(
        left = Side(border_style = 'thin', color = 'FF000000'),
        right = Side(border_style = 'thin', color = 'FF000000'),
        top = Side(border_style = 'thin', color = 'FF000000'),
        bottom = Side(border_style = 'thin', color = 'FF000000')),
    'fill': PatternFill(),
    'alignment': Alignment(horizontal = 'general', vertical = 'center')}
    
FSTYLE = {'font': Font(name = 'Arial', size = 10, italic = True),
    'border': Border(),
    'fill': PatternFill(),
    'alignment': Alignment(horizontal = 'general', vertical = 'center')}


class ExcelReport(openpyxl.Workbook):
    '''
    класс для генерации отчетов в формате книги Excel
    '''
    def worksheet_as_page(self, mdlexample, name ='_', 
        hstyle = HSTYLE, bstyle = BSTYLE, fstyle = FSTYLE, note = ''):
        '''
        создание отчета экземпляра модели в виде листа
            mdlexample - экземпляр класса Model или унаследованного от Model
            name - строка, имя листа
            hstyle, bstyle, fstyle - словари следующего 
            содержания: {font: объект класса Font, Border: ..., 
            fill: PatternFill, alignment: Alignment}
            note - примечание строка, по умолчанию пишется в footer отчета
            добавляет worksheet в книгу и возвращает ссылку на лист
        '''
        ws = self.create_sheet(name)
        if mdlexample == None: 
            ws.append(['the model no data'])
        else:
            if isinstance(mdlexample, models.Model):
                fieldlist = mdlexample._meta.get_fields()
                for fld in fieldlist:
                    if not (isinstance(fld, models.fields.related.RelatedField) or isinstance(fld, models.fields.reverse_related.ManyToOneRel)):
                        ws.append([fld.verbose_name.upper(), eval('mdlexample.' + fld.name)])
                for row in ws.rows:
                    for cell in row:
                        cell.font = HSTYLE['font']
                        cell.border = HSTYLE['border']
                        if cell.column == 1:
                            cell.fill = HSTYLE['fill']
        return ws

    def worksheet_as_list(self, mdlset, name ='_', 
        hstyle = HSTYLE, bstyle = BSTYLE, fstyle = FSTYLE, note = ''):
        '''
        создание отчета экземпляра модели в виде списка - таблицы
            mdlset - набор (QuerySet) экземпляров класса Model или унаследованного от Model
            name - строка, имя листа
            hstyle, bstyle, fstyle - словари следующего 
            содержания: {font: объект класса Font, Border: ..., 
            fill: PatternFill, alignment: Alignment}
            note - примечание строка, по умолчанию пишется в footer отчета
            добавляет worksheet в книгу и возвращает ссылку на лист
        '''
        ws = self.create_sheet(name)
        if mdlset == None: 
            ws.append(['the model no data'])
        else:
            if len(mdlset) == 0:
                ws.append(['the model no data'])
            else:
                if isinstance(mdlset[0], models.Model):
                    fieldlist = mdlset[0]._meta.get_fields()
                    fieldnamelist = []
                    header = []
                    for fld in fieldlist:
                        if not (isinstance(fld, models.fields.related.RelatedField) 
                            or isinstance(fld, models.fields.reverse_related.ManyToOneRel)
                            or isinstance(fld, models.fields.reverse_related.ManyToManyRel)
                            or isinstance(fld, models.FileField)):
                            fieldnamelist.append(fld.name)
                            header.append(fld.verbose_name.upper())
                    ws.append(header)        
                    for mdlexample in mdlset:
                        datarow = []
                        for fldname in fieldnamelist:
                            datarow.append(eval('mdlexample.' + fldname))
                        ws.append(datarow)
                    for row in ws.rows:
                        for cell in row:
                            if cell.row == 1:
                                cell.font = HSTYLE['font']
                                cell.border = HSTYLE['border']
                                cell.fill = HSTYLE['fill']
                            else:
                                cell.font = BSTYLE['font']
                                cell.border = BSTYLE['border']
                                cell.fill = BSTYLE['fill']
        return ws

    def worksheet_as_2list(self, mdlset, relname = '', name ='_', 
        hstyle = HSTYLE, bstyle = BSTYLE, fstyle = FSTYLE, note = ''):
        '''
        создание отчета экземпляра модели в виде сложного списка - таблицы модели и сязанной модели
            mdlset - набор (QuerySet) экземпляров класса Model или унаследованного от Model
            name - строка, имя листа
            relname - имя поля связанной модели
            hstyle, bstyle, fstyle - словари следующего 
            содержания: {font: объект класса Font, Border: ..., 
            fill: PatternFill, alignment: Alignment}
            note - примечание строка, по умолчанию пишется в footer отчета
            добавляет worksheet в книгу и возвращает ссылку на лист
        '''
        ws = self.create_sheet(name)
        if mdlset == None: 
            ws.append(['the model no data'])
            return ws
        if len(mdlset) == 0:
            ws.append(['the model no data'])
            return ws
        if not isinstance(mdlset[0], models.Model):
            ws.append(['the report does not match the model type'])
            return ws
# формируем заголовок и список полей связанной модели
        relset = eval('mdlset[0].' + relname + '_set.all()')
        fieldnamelist = []
        if len(relset) == 0:
            ws.append(['rel model no data'])
        else:
            if not isinstance(mdlset[0], models.Model):
                ws.append(['the report does not match the rel model type'])
            else:
                fieldlist = relset[0]._meta.get_fields()
                header = []
                for fld in fieldlist:
                    if not (isinstance(fld, models.fields.related.RelatedField) 
                        or isinstance(fld, models.fields.reverse_related.ManyToOneRel)
                        or isinstance(fld, models.fields.reverse_related.ManyToManyRel)
                        or isinstance(fld, models.FileField)):
                        fieldnamelist.append(fld.name)
                        header.append(fld.verbose_name.upper())
                ws.append(header)        
# перебираем экземпляры модели            
        for mdlexample in mdlset:
            ws.append([mdlexample.__str__()])
# получаем набор (QuerySet) экземпляров связанной модели
            relset = eval('mdlexample.' + relname + '_set.all()')
            if len(relset) == 0:
                datarow = []
                for fldname in fieldnamelist: datarow.append(' - ')
            else:
                for relmdlexample in relset: 
# получаем данные экземпляра связанной модели для полей из списка
                    datarow = []
                    for fldname in fieldnamelist:
                        datarow.append(eval('relmdlexample.' + fldname))
            ws.append(datarow)
# формируем                    
            for row in ws.rows:
                for cell in row:
                    if cell.row == 1:
                        cell.font = HSTYLE['font']
                        cell.border = HSTYLE['border']
                        cell.fill = HSTYLE['fill']
                    else:
                        cell.font = BSTYLE['font']
                        cell.border = BSTYLE['border']
                        cell.fill = BSTYLE['fill']
        return ws
